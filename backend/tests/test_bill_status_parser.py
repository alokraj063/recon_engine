"""
IREPS "Bill Status" parsing — CURRENT tabular export.

Fixtures are built in code: a header row of the 20 IREPS columns (any
order, any starting column/row) plus data rows underneath. The parser
must locate columns by their header TEXT, never by position, so several
tests deliberately shift the header row/column or reorder it.

An optional smoke test at the bottom runs against the real sample export
(gitignored) when present, purely as a sanity check — the structural
guarantees above are the ones that must hold for the next dated file too.
"""

import datetime
import sys
from pathlib import Path

import openpyxl
import pandas as pd
import pytest
import xlwt

BACKEND = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND))

from recon.parsers import parse_bill_status  # noqa: E402
from recon.parsers.bill_status import (BillStatusFormatError,  # noqa: E402
                                       HEADER_TO_FIELD)
from recon.sources import get_adapter  # noqa: E402
from recon.sources.base import SelfCheckError  # noqa: E402
from recon.sources.ireps_bills import _parse_recovery_details  # noqa: E402

HEADERS = list(HEADER_TO_FIELD)  # source column order, as IREPS exports it

ROW = {
    "Contract No": "2251424102381",
    "Contract Date": pd.Timestamp("2025-12-22").to_pydatetime(),
    "Bill Date": pd.Timestamp("2026-07-23").to_pydatetime(),
    "Bill Number": "1331000197",
    "Zone": "ICF",
    "Party Name": "SOME VENDOR PRIVATE LIMITED-H",
    "PartyCode": "MM23:833",
    "CO6 No": 13010326014461,
    "CO6 Date": pd.Timestamp("2026-07-27").to_pydatetime(),
    "Status": "PAYMENT MADE",
    "Bill Amt": 365753,
    "Passed Amt": 365752.8,
    "Deducted Amt": 6509.2,
    "Net Amt": 359243,
    "CO7 No": 13010326701086,
    "CO7 Date": pd.Timestamp("2026-07-28").to_pydatetime(),
    "Payment Advice Date": pd.Timestamp("2026-07-28").to_pydatetime(),
    "Accounting Unit": "ICF HEADQUARTER",
    "Reason For Return": None,
    "Recovery Details": "GST TDS DEDUCTION: 6199.2 INCOME TAX - CONTR(Section-194Q): 310",
}


def workbook(tmp_path, name="bills.xlsx", sheet="bill",
            headers=HEADERS, header_row=1, header_col=1):
    """A tabular Bill Status workbook: header row, no data yet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    for i, h in enumerate(headers):
        ws.cell(row=header_row, column=header_col + i, value=h)
    return wb, ws, tmp_path / name


def write_row(ws, row_idx, values=ROW, headers=HEADERS, header_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row_idx, column=header_col + i, value=values.get(h))


def save(wb, path):
    wb.save(path)
    return path


# --- legacy .xls (BIFF) fixtures -----------------------------------------
# Same shape as the openpyxl helpers above, built with xlwt instead, so
# every .xls-specific test exercises the real xlrd code path rather than
# reusing an openpyxl-derived file under a renamed extension.

_XLS_DATE_STYLE = xlwt.easyxf(num_format_str="DD/MM/YYYY")


def workbook_xls(tmp_path, name="bills.xls", sheet="bill",
                 headers=HEADERS, header_row=1, header_col=1):
    wb = xlwt.Workbook()
    ws = wb.add_sheet(sheet)
    for i, h in enumerate(headers):
        ws.write(header_row - 1, header_col - 1 + i, h)
    return wb, ws, tmp_path / name


def write_row_xls(ws, row_idx, values=ROW, headers=HEADERS, header_col=1):
    for i, h in enumerate(headers):
        v = values.get(h)
        if v is None:
            continue
        col = header_col - 1 + i
        if isinstance(v, datetime.datetime):
            ws.write(row_idx - 1, col, v, _XLS_DATE_STYLE)
        else:
            ws.write(row_idx - 1, col, v)


# --- Test 1 / Test 2 — current tabular format + header mapping ----------

def test_header_row_maps_every_column(tmp_path):
    wb, ws, path = workbook(tmp_path)
    write_row(ws, 2)
    df = parse_bill_status(save(wb, path))

    assert len(df) == 1
    row = df.iloc[0]
    assert row["ContractNo"] == "2251424102381"
    assert row["ContractDate"] == pd.Timestamp("2025-12-22")
    assert row["BillDate"] == pd.Timestamp("2026-07-23")
    assert row["BillNumber"] == "1331000197"
    assert row["Zone"] == "ICF"
    assert row["PartyName"] == "SOME VENDOR PRIVATE LIMITED-H"
    assert row["PartyCode"] == "MM23:833"
    assert row["CO6No"] == 13010326014461
    assert row["CO6Date"] == pd.Timestamp("2026-07-27")
    assert row["Status"] == "PAYMENT MADE"
    assert row["BillAmt"] == 365753
    assert row["PassedAmt"] == 365752.8
    assert row["DeductedAmt"] == 6509.2
    assert row["NetAmt"] == 359243
    assert row["CO7No"] == 13010326701086
    assert row["CO7Date"] == pd.Timestamp("2026-07-28")
    assert row["PaymentAdviceDate"] == pd.Timestamp("2026-07-28")
    assert row["AccountingUnit"] == "ICF HEADQUARTER"
    assert row["ReasonForReturn"] is None
    assert row["RecoveryDetails"] == ROW["Recovery Details"]


def test_multiple_rows_extracted(tmp_path):
    wb, ws, path = workbook(tmp_path)
    for i, bill_no in enumerate(("A", "B", "C")):
        write_row(ws, 2 + i, {**ROW, "Bill Number": bill_no})
    df = parse_bill_status(save(wb, path))

    assert list(df["BillNumber"]) == ["A", "B", "C"]
    assert list(df["DataRow"]) == [2, 3, 4]


# --- Test 3/4/5 — status values -------------------------------------

def test_payment_made_row(tmp_path):
    wb, ws, path = workbook(tmp_path)
    write_row(ws, 2, {**ROW, "Status": "PAYMENT MADE"})
    df = parse_bill_status(save(wb, path))
    assert df.iloc[0]["Status"] == "PAYMENT MADE"


def test_registered_row_with_co7_not_issued(tmp_path):
    wb, ws, path = workbook(tmp_path)
    write_row(ws, 2, {**ROW, "Status": "REGISTERED",
                      "CO7 No": "----", "CO7 Date": "----",
                      "Payment Advice Date": "NA"})
    df = parse_bill_status(save(wb, path))
    row = df.iloc[0]
    assert row["Status"] == "REGISTERED"
    assert row["CO7No"] is None
    assert pd.isna(row["CO7Date"])
    assert pd.isna(row["PaymentAdviceDate"])


def test_returned_row_carries_reason(tmp_path):
    reason = ("#Invoice number mismatch. The invoice number has been "
             "wrongly entered in the bill as 106100033 instead of "
             "1061000033.")
    wb, ws, path = workbook(tmp_path)
    write_row(ws, 2, {**ROW, "Status": "RETURNED",
                      "Reason For Return": reason,
                      "CO7 No": "----", "CO7 Date": "----"})
    df = parse_bill_status(save(wb, path))
    row = df.iloc[0]
    assert row["Status"] == "RETURNED"
    assert row["ReasonForReturn"] == reason


# --- Test 6/7 — Recovery Details ------------------------------------

def test_recovery_details_string_preserved_verbatim(tmp_path):
    text = ("DEPOSIT STORES-LIQUIDITY DAMAGE: 106200 GST TDS DEDUCTION: "
           "18000 INCOME TAX - CONTR(Section-194Q): 900")
    wb, ws, path = workbook(tmp_path)
    write_row(ws, 2, {**ROW, "Recovery Details": text})
    df = parse_bill_status(save(wb, path))
    assert df.iloc[0]["RecoveryDetails"] == text


def test_no_recovery_details(tmp_path):
    wb, ws, path = workbook(tmp_path)
    write_row(ws, 2, {**ROW, "Recovery Details": None})
    df = parse_bill_status(save(wb, path))
    assert pd.isna(df.iloc[0]["RecoveryDetails"]) or df.iloc[0]["RecoveryDetails"] is None


def test_recovery_details_splits_into_head_amount_pairs():
    text = ("DEPOSIT STORES-LIQUIDITY DAMAGE: 106200 GST TDS DEDUCTION: "
           "18000 INCOME TAX - CONTR(Section-194Q): 900")
    items = _parse_recovery_details(text)
    assert items == [
        ("DEPOSIT STORES-LIQUIDITY DAMAGE", "106200"),
        ("GST TDS DEDUCTION", "18000"),
        ("INCOME TAX - CONTR(Section-194Q)", "900"),
    ]
    assert _parse_recovery_details(None) == []


def test_recovery_details_with_leading_dot_amount():
    """'OTHER CHARGES: .31' — an amount with no leading digit."""
    items = _parse_recovery_details("OTHER CHARGES: .31")
    assert items == [("OTHER CHARGES", ".31")]


# --- Test 8 — CO7 placeholder ----------------------------------------

def test_co7_dashes_become_null_amounts_and_dates_untouched(tmp_path):
    wb, ws, path = workbook(tmp_path)
    write_row(ws, 2, {**ROW, "CO7 No": "----", "CO7 Date": "----"})
    df = parse_bill_status(save(wb, path))
    row = df.iloc[0]
    assert row["CO7No"] is None
    assert pd.isna(row["CO7Date"])
    # unrelated fields on the same row are untouched
    assert row["BillAmt"] == 365753
    assert row["Status"] == "PAYMENT MADE"


# --- Test 9 — identifier preservation ---------------------------------

def test_identifiers_are_not_coerced_to_numbers(tmp_path):
    wb, ws, path = workbook(tmp_path)
    write_row(ws, 2, {**ROW, "Contract No": "0606WC18000068",
                      "Bill Number": "-", "PartyCode": "MM23:833"})
    df = parse_bill_status(save(wb, path))
    row = df.iloc[0]
    assert row["ContractNo"] == "0606WC18000068"
    assert row["BillNumber"] == "-"          # preserved verbatim, not nulled
    assert row["PartyCode"] == "MM23:833"
    for col in ("BillAmt", "PassedAmt", "DeductedAmt", "NetAmt"):
        assert pd.api.types.is_numeric_dtype(df[col])
    for col in ("ContractNo", "BillNumber", "PartyCode"):
        assert not pd.api.types.is_numeric_dtype(df[col])


def test_numeric_looking_identifiers_stay_exact(tmp_path):
    """CO6No/CO7No/BillNumber are reference numbers, not amounts — no
    rounding, no float artifacts."""
    wb, ws, path = workbook(tmp_path)
    write_row(ws, 2, {**ROW, "Bill Number": "1061000033",
                      "CO6 No": 13010326014461, "CO7 No": 13010326701086})
    df = parse_bill_status(save(wb, path))
    row = df.iloc[0]
    assert row["BillNumber"] == "1061000033"
    assert row["CO6No"] == 13010326014461
    assert row["CO7No"] == 13010326701086


# --- legacy .xls (BIFF) support ------------------------------------------

def test_xls_format_is_parsed(tmp_path):
    wb, ws, path = workbook_xls(tmp_path)
    write_row_xls(ws, 2)
    df = parse_bill_status(save(wb, path))

    assert len(df) == 1
    row = df.iloc[0]
    assert row["BillNumber"] == "1331000197"
    assert row["Zone"] == "ICF"
    assert row["Status"] == "PAYMENT MADE"
    assert row["NetAmt"] == 359243
    assert row["RecoveryDetails"] == ROW["Recovery Details"]


def test_xls_and_xlsx_produce_identical_records(tmp_path):
    """Same content, two containers — the Silver record must not depend
    on which one was uploaded."""
    wbx, wsx, pathx = workbook(tmp_path, name="a.xlsx")
    write_row(wsx, 2)
    dfx = parse_bill_status(save(wbx, pathx))

    wbl, wsl, pathl = workbook_xls(tmp_path, name="a.xls")
    write_row_xls(wsl, 2)
    dfl = parse_bill_status(save(wbl, pathl))

    pd.testing.assert_frame_equal(
        dfx.drop(columns=["Sheet"]), dfl.drop(columns=["Sheet"]))


def test_xls_dates_convert_and_identifiers_stay_exact(tmp_path):
    wb, ws, path = workbook_xls(tmp_path)
    write_row_xls(ws, 2)
    df = parse_bill_status(save(wb, path))
    row = df.iloc[0]

    assert row["BillDate"] == pd.Timestamp("2026-07-23")
    assert row["CO6Date"] == pd.Timestamp("2026-07-27")
    # BIFF has no integer cell type (every number is a float) — CO6No
    # must come back exact, not 13010326014461.0000001-style float noise
    assert row["CO6No"] == 13010326014461
    assert float(row["CO6No"]).is_integer()


def test_xls_co7_placeholder_and_multiple_sheets(tmp_path):
    wb = xlwt.Workbook()
    for sheet, bill_no in (("Friction", "F1"), ("Rohtak", "R1")):
        ws = wb.add_sheet(sheet)
        for i, h in enumerate(HEADERS):
            ws.write(0, i, h)
        write_row_xls(ws, 2, {**ROW, "Bill Number": bill_no,
                              "CO7 No": "----", "CO7 Date": "----"})
    path = save(wb, tmp_path / "multi.xls")
    df = parse_bill_status(path)

    assert set(df["Sheet"]) == {"Friction", "Rohtak"}
    assert set(df["BillNumber"]) == {"F1", "R1"}
    assert df["CO7No"].isna().all()


def test_xls_adapter_selfcheck(tmp_path):
    wb, ws, path = workbook_xls(tmp_path)
    write_row_xls(ws, 2)
    path = save(wb, path)

    adapter = get_adapter("bill_status", "ireps")
    assert ".xls" in adapter.file_kinds
    gold = adapter.to_gold(adapter.parse(path, {}), {})
    assert adapter.selfcheck(gold, path, {})["parsed_count"] == 1


# --- the column list is a moving target ---------------------------------
# This export already changed shape once. A future one that adds or drops
# a column must keep ingesting: only REQUIRED_HEADERS are load-bearing.

def test_an_added_column_reaches_silver_under_its_own_header(tmp_path):
    """A column IREPS adds that no map knows: extracted anyway, named by
    the header text the source itself wrote."""
    headers = HEADERS + ["UTR No", "Payment Date"]
    wb, ws, path = workbook(tmp_path, headers=headers)
    write_row(ws, 2, {**ROW, "UTR No": "HSBCN26073100123",
                      "Payment Date": pd.Timestamp("2026-07-31").to_pydatetime()},
              headers=headers)
    df = parse_bill_status(save(wb, path))

    assert df.iloc[0]["UTR No"] == "HSBCN26073100123"
    assert df.iloc[0]["Payment Date"] == datetime.datetime(2026, 7, 31)
    # the known columns are unaffected by the newcomers
    assert df.iloc[0]["BillNumber"] == "1331000197"
    assert df.iloc[0]["NetAmt"] == 359243


def test_an_added_column_rides_through_to_gold_extras(tmp_path):
    """...and reaches gold without an adapter change: unmapped columns
    survive to_gold, which is what puts them in gold.bills.extras."""
    headers = HEADERS + ["UTR No"]
    wb, ws, path = workbook(tmp_path, headers=headers)
    write_row(ws, 2, {**ROW, "UTR No": "HSBCN26073100123"}, headers=headers)
    path = save(wb, path)

    adapter = get_adapter("bill_status", "ireps")
    gold = adapter.to_gold(adapter.parse(path, {}), {})["bills"]
    assert gold.iloc[0]["UTR No"] == "HSBCN26073100123"
    assert gold.iloc[0]["net_payable_amount"] == 359243


def test_a_dropped_optional_column_still_parses(tmp_path):
    """An export that stops sending 'Accounting Unit' / 'Reason For
    Return' parses fine — those fields simply come back empty."""
    headers = [h for h in HEADERS
               if h not in ("Accounting Unit", "Reason For Return")]
    wb, ws, path = workbook(tmp_path, headers=headers)
    write_row(ws, 2, headers=headers)
    df = parse_bill_status(save(wb, path))

    assert len(df) == 1
    assert pd.isna(df.iloc[0]["AccountingUnit"])
    assert pd.isna(df.iloc[0]["ReasonForReturn"])
    assert df.iloc[0]["BillNumber"] == "1331000197"
    # and the gold transform still runs over the NA-filled columns
    adapter = get_adapter("bill_status", "ireps")
    gold = adapter.to_gold(adapter.parse(save(wb, path), {}), {})["bills"]
    assert pd.isna(gold.iloc[0]["org_unit"])
    assert gold.iloc[0]["recovery_count"] == 2


def test_a_missing_required_column_fails_loud_and_names_it(tmp_path):
    """Dropping a load-bearing column is NOT tolerated — a bill with no
    net amount cannot be reconciled, so the upload must fail, saying so."""
    headers = [h for h in HEADERS if h != "Net Amt"]
    wb, ws, path = workbook(tmp_path, headers=headers)
    write_row(ws, 2, headers=headers)

    with pytest.raises(BillStatusFormatError) as err:
        parse_bill_status(save(wb, path))
    assert "Net Amt" in str(err.value)


def test_sheets_with_different_extra_columns_share_one_frame(tmp_path):
    """Two sheets, each with its own added column: every row keeps its
    own value and misses the other's."""
    wb = openpyxl.Workbook()
    first = wb.active
    first.title = "Friction"
    second = wb.create_sheet("Hosur")
    for ws, extra, bill in ((first, "UTR No", "F1"), (second, "Batch", "H1")):
        headers = HEADERS + [extra]
        for i, h in enumerate(headers):
            ws.cell(row=1, column=1 + i, value=h)
        write_row(ws, 2, {**ROW, "Bill Number": bill, extra: f"{extra}-val"},
                  headers=headers)
    df = parse_bill_status(save(wb, tmp_path / "two_sheets.xlsx"))

    assert set(df["BillNumber"]) == {"F1", "H1"}
    by_bill = df.set_index("BillNumber")
    assert by_bill.loc["F1", "UTR No"] == "UTR No-val"
    assert by_bill.loc["H1", "Batch"] == "Batch-val"
    assert pd.isna(by_bill.loc["F1", "Batch"])
    assert pd.isna(by_bill.loc["H1", "UTR No"])


# --- Test 10 — invalid file --------------------------------------------

def test_missing_required_columns_raises(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1, value="Field")
    ws.cell(row=1, column=2, value="Value")
    ws.cell(row=2, column=1, value="Bill No.")
    ws.cell(row=2, column=2, value="BILL/2026/SR/00125")
    path = save(wb, tmp_path / "not_bill_status.xlsx")

    with pytest.raises(BillStatusFormatError):
        parse_bill_status(path)


def test_empty_workbook_raises(tmp_path):
    wb = openpyxl.Workbook()
    path = save(wb, tmp_path / "empty.xlsx")
    with pytest.raises(BillStatusFormatError):
        parse_bill_status(path)


# --- Test 11 — position independence / similar future files ------------

def test_header_at_different_row_and_column(tmp_path):
    """The header need not start at A1 — only its text is located."""
    wb, ws, path = workbook(tmp_path, header_row=3, header_col=2)
    write_row(ws, 4, header_col=2)
    df = parse_bill_status(save(wb, path))
    assert len(df) == 1
    assert df.iloc[0]["BillNumber"] == "1331000197"


def test_header_columns_reordered(tmp_path):
    """Column order in the export must not matter."""
    shuffled = list(reversed(HEADERS))
    wb, ws, path = workbook(tmp_path, headers=shuffled)
    write_row(ws, 2, headers=shuffled)
    df = parse_bill_status(save(wb, path))
    assert len(df) == 1
    assert df.iloc[0]["BillNumber"] == "1331000197"
    assert df.iloc[0]["NetAmt"] == 359243


def test_a_sibling_dated_file_with_different_row_count(tmp_path):
    """A different filename/date, more rows, different values — nothing
    hardcoded should care."""
    wb, ws, path = workbook(tmp_path, name="BILL STATUS 15082026.xlsx")
    for i in range(7):
        write_row(ws, 2 + i, {**ROW, "Bill Number": f"BILL-{i}",
                              "Net Amt": 1000 + i})
    df = parse_bill_status(save(wb, path))
    assert len(df) == 7
    assert list(df["BillNumber"]) == [f"BILL-{i}" for i in range(7)]
    assert list(df["NetAmt"]) == [1000 + i for i in range(7)]


def test_only_a_matching_sheet_is_processed(tmp_path):
    """A workbook with an unrelated sheet alongside the real one: only
    the sheet carrying the header is read."""
    wb, ws, path = workbook(tmp_path, sheet="bill")
    write_row(ws, 2)
    other = wb.create_sheet("filters")
    other.cell(row=1, column=1, value="Railway Zone")
    other.cell(row=2, column=1, value="ICF")
    df = parse_bill_status(save(wb, path))
    assert len(df) == 1
    assert df.iloc[0]["Sheet"] == "bill"


def test_blank_rows_between_data_are_skipped(tmp_path):
    wb, ws, path = workbook(tmp_path)
    write_row(ws, 2, {**ROW, "Bill Number": "FIRST"})
    write_row(ws, 4, {**ROW, "Bill Number": "SECOND"})   # row 3 left blank
    df = parse_bill_status(save(wb, path))
    assert list(df["BillNumber"]) == ["FIRST", "SECOND"]


# --- adapter seam --------------------------------------------------------

def test_adapter_wraps_the_parser_output_unchanged(tmp_path):
    wb, ws, path = workbook(tmp_path)
    write_row(ws, 2)
    path = save(wb, path)

    adapter = get_adapter("bill_status", "ireps")
    silver = adapter.parse(path, {})
    assert set(silver.frames) == {"bills"}
    assert "BillNumber" in silver.frames["bills"].columns
    assert "bill_number" not in silver.frames["bills"].columns
    assert "RecoveryDetails" in silver.frames["bills"].columns
    # gold-schema derivations are NOT in silver
    assert "Recoveries" not in silver.frames["bills"].columns

    gold = adapter.to_gold(silver, {})
    bill = gold["bills"].iloc[0]
    assert bill["bill_number"] == "1331000197"
    assert bill["net_payable_amount"] == 359243
    assert bill["recovery_count"] == 2
    assert bill["recoveries"] == {"GST TDS DEDUCTION": "6199.2",
                                  "INCOME TAX - CONTR(Section-194Q)": "310"}
    assert bill["recovery_sum"] == pytest.approx(6509.2)
    assert bool(bill["recovery_check"]) is True
    assert bool(bill["net_check"]) is True
    assert pd.isna(bill["header_row"])       # obsolete for this format
    assert pd.isna(bill["unparsed_header"])

    recs = gold["recoveries"]
    assert len(recs) == 2
    assert set(recs["recovery_head"]) == {"GST TDS DEDUCTION",
                                          "INCOME TAX - CONTR(Section-194Q)"}
    assert adapter.selfcheck(gold, path, {})["parsed_count"] == 1


def test_adapter_parse_fails_loud_on_a_non_bill_status_workbook(tmp_path):
    wb = openpyxl.Workbook()
    path = save(wb, tmp_path / "wrong.xlsx")
    with pytest.raises(BillStatusFormatError):
        get_adapter("bill_status", "ireps").parse(path, {})


def test_adapter_selfcheck_fails_loud_on_a_header_only_workbook(tmp_path):
    """Header row matches, but zero bill rows underneath — the workbook
    parses without error, so it's selfcheck()'s job to catch this."""
    wb, _ws, path = workbook(tmp_path)
    path = save(wb, path)
    adapter = get_adapter("bill_status", "ireps")
    gold = adapter.to_gold(adapter.parse(path, {}), {})
    with pytest.raises(SelfCheckError):
        adapter.selfcheck(gold, path, {})


# --- optional: the real export ------------------------------------------

SAMPLE_DIR = (BACKEND.parent
              / "Receipt_reconciliation_and_IDR _ Requested_sample_documents")


def _sample():
    for p in sorted(SAMPLE_DIR.glob("BILL STATUS*.xlsx")):
        if not p.name.startswith("~$"):
            return p
    return None


@pytest.mark.skipif(_sample() is None, reason="sample documents not present")
def test_real_export_smoke():
    df = parse_bill_status(_sample())

    assert len(df) > 0
    assert set(HEADER_TO_FIELD.values()) <= set(df.columns)
    for col in ("BillAmt", "PassedAmt", "DeductedAmt", "NetAmt"):
        assert pd.api.types.is_numeric_dtype(df[col])
    # ContractNo/BillNumber mix real identifiers with non-numeric
    # placeholders ('-', 'NA') in the real export, so they must never be
    # coerced to a numeric dtype; CO6No/CO7No happen to be uniformly
    # numeric in this file (pandas infers int64 on its own) but must
    # never gain a float artifact (identifiers, not amounts)
    assert not pd.api.types.is_numeric_dtype(df["ContractNo"])
    assert not pd.api.types.is_numeric_dtype(df["BillNumber"])
    for col in ("CO6No", "CO7No"):
        non_null = df[col].dropna()
        assert (non_null == non_null.astype("int64")).all()
    returned = df[df["Status"] == "RETURNED"]
    assert returned["ReasonForReturn"].notna().any()
