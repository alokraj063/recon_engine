"""
Excel output. Nothing here decides anything; it only lays out frames the
engine already produced.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
BANK_FILL = PatternFill("solid", fgColor="FCE4D6")
BILL_FILL = PatternFill("solid", fgColor="DDEBF7")
BODY_FONT = Font(name="Arial", size=10)
HEAD_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")


def write_workbook(out, path):
    """Four sheets: summary, matched, the combined exception queue, and
    the deduction detail."""
    with pd.ExcelWriter(path, engine="openpyxl") as xl:
        out["summary"].to_excel(xl, sheet_name="Summary", index=False)
        (out["matched"].drop(columns=["bill_indices", "candidate_indices"],
                             errors="ignore")
            .to_excel(xl, sheet_name="Matched", index=False))
        # Candidates is structured (list of dicts) for the API; Excel gets
        # the flat CandidateSummary string instead.
        (out["queue"].drop(columns=["Candidates", "bill_indices",
                                    "candidate_indices"], errors="ignore")
            .to_excel(xl, sheet_name="Exception_Queue", index=False))
        if "recoveries" in out:
            out["recoveries"].to_excel(xl, sheet_name="Recovery_Detail", index=False)
    _format(path)
    return path


def _format(path):
    wb = load_workbook(path)
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = HEAD_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            width = max((len(str(c.value)) for c in col[:60] if c.value), default=8)
            ws.column_dimensions[letter].width = min(max(width + 2, 10), 42)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = BODY_FONT

    # Colour the two sides of the queue so they read apart at a glance.
    ws = wb["Exception_Queue"]
    for row in ws.iter_rows(min_row=2):
        row[0].fill = BANK_FILL if row[0].value == "BANK_ONLY" else BILL_FILL
    wb.save(path)
