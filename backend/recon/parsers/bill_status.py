"""
Parser for the IREPS "Bill Status" export.

Current format (since 2026-08): a plain table — one header row naming the
20 IREPS columns, one data row per bill underneath:

    Contract No | Contract Date | Bill Date | Bill Number | Zone |
    Party Name | PartyCode | CO6 No | CO6 Date | Status | Bill Amt |
    Passed Amt | Deducted Amt | Net Amt | CO7 No | CO7 Date |
    Payment Advice Date | Accounting Unit | Reason For Return |
    Recovery Details

Columns are located by their HEADER TEXT, never by a fixed row or column
index, so the header row may sit at the top of any worksheet — the parser
scans every sheet and reads whichever ones carry it, and raises
BillStatusFormatError if none do (a file that isn't this format must fail
loud, not silently ingest as zero bills). `Recovery Details` is ONE
free-text cell per bill ('<head>: <amt> <head>: <amt> ...'); splitting it
into structured recovery lines is a Gold-layer concern (see
sources/ireps_bills.py), not this parser's — Silver keeps it verbatim.

The column list is treated as a MOVING TARGET, because it is one: this
export already changed shape once. Only REQUIRED_HEADERS have to be
present for a sheet to count as Bill Status — the four without which a
bill cannot be identified, registered, stated or paid. Everything else is
optional and NA-fills when absent, and any column IREPS adds that this
map doesn't know is still extracted, under its own header text as the
silver field name (see `_find_header`). So a future export that adds
'UTR No' or drops 'Accounting Unit' keeps ingesting: the new field
reaches gold through the adapter's `extras`, and only a genuinely
unusable file fails — with a message naming the headers it missed.

IREPS exports the same tabular format as either `.xlsx`/`.xlsm` (OOXML,
read via openpyxl) or the legacy `.xls` (BIFF, read via xlrd) — both are
accepted. `_open_workbook` picks the reader by extension and hands back
the same tiny surface (`.sheetnames`, `wb[name].max_row`,
`wb[name].iter_rows(...)`) either way, so everything below this point is
format-agnostic. The one real difference between the two libraries: BIFF
has no integer cell type (every number is a float) and xlrd resolves a
date cell to an Excel serial number rather than a datetime — `_XlsSheet`
normalizes both so a whole-number reference (`CO6 No`) and a date cell
come back exactly as openpyxl would already give them.
"""

import re
from pathlib import Path
from typing import NamedTuple, Optional

import openpyxl
import pandas as pd
import xlrd

# source header text -> silver field name (source-native vocabulary)
HEADER_TO_FIELD = {
    "Contract No": "ContractNo",
    "Contract Date": "ContractDate",
    "Bill Date": "BillDate",
    "Bill Number": "BillNumber",
    "Zone": "Zone",
    "Party Name": "PartyName",
    "PartyCode": "PartyCode",
    "CO6 No": "CO6No",
    "CO6 Date": "CO6Date",
    "Status": "Status",
    "Bill Amt": "BillAmt",
    "Passed Amt": "PassedAmt",
    "Deducted Amt": "DeductedAmt",
    "Net Amt": "NetAmt",
    "CO7 No": "CO7No",
    "CO7 Date": "CO7Date",
    "Payment Advice Date": "PaymentAdviceDate",
    "Accounting Unit": "AccountingUnit",
    "Reason For Return": "ReasonForReturn",
    "Recovery Details": "RecoveryDetails",
}

# The columns that MUST be present for a worksheet to be a Bill Status
# table at all: the bill's own number, the registration reference it is
# upserted by, its state, and the amount that has to reach the bank.
# Every other column above is optional — a missing one NA-fills rather
# than failing the upload.
REQUIRED_HEADERS = ("Bill Number", "CO6 No", "Status", "Net Amt")

# Silver-layer bookkeeping, not one of the 20 IREPS columns: which sheet a
# row came from and its physical Excel row. Kept for the same reason the
# gold schema already declares slots for them (`sheet`, `data_row`) —
# `data_row` in particular is a real dependency of
# engine.group_bill_attempts' resubmission tie-break ("a smaller row is
# the more recent attempt"), not decoration.
BOOKKEEPING_FIELDS = ["Sheet", "DataRow"]
FIELDS = list(HEADER_TO_FIELD.values()) + BOOKKEEPING_FIELDS

DATE_FIELDS = ["ContractDate", "BillDate", "CO6Date", "CO7Date",
              "PaymentAdviceDate"]
AMOUNT_FIELDS = ["BillAmt", "PassedAmt", "DeductedAmt", "NetAmt"]
# identifiers: never coerced to numeric, kept exactly as the source wrote
# them (a bill/contract/reference number is not an amount)
IDENTIFIER_FIELDS = ["ContractNo", "BillNumber", "CO6No", "CO7No"]

MAX_HEADER_SCAN_ROWS = 5
# '----' is IREPS's placeholder for "CO7 not issued yet" — the only
# identifier field where the export itself uses a placeholder rather than
# leaving the cell blank
_DASH_PLACEHOLDER_RE = re.compile(r"^-{2,}$")


class BillStatusFormatError(ValueError):
    """The workbook carries no Bill Status header row in any worksheet."""


class _XlsSheet:
    """Adapts one xlrd sheet to the small openpyxl.Worksheet surface this
    parser uses (`max_row`, `iter_rows(min_row, max_row, values_only)`),
    normalizing cell values to the same plain Python types openpyxl
    already gives: a whole-number cell -> int (BIFF has no int type of
    its own, every number is a float), a date cell -> datetime, a blank
    cell -> None."""

    def __init__(self, xlrd_sheet, datemode):
        self._sheet = xlrd_sheet
        self._datemode = datemode
        self.max_row = xlrd_sheet.nrows

    def iter_rows(self, min_row=1, max_row=None, values_only=True):
        max_row = self.max_row if max_row is None else min(max_row, self.max_row)
        for r in range(min_row - 1, max_row):
            yield tuple(self._cell(r, c) for c in range(self._sheet.ncols))

    def _cell(self, r, c):
        cell = self._sheet.cell(r, c)
        if cell.ctype == xlrd.XL_CELL_DATE:
            try:
                return xlrd.xldate.xldate_as_datetime(cell.value, self._datemode)
            except (xlrd.xldate.XLDateError, ValueError):
                return cell.value
        if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
            return None
        v = cell.value
        if isinstance(v, float) and v.is_integer():
            return int(v)
        return v if v != "" else None


class _XlsWorkbook:
    """Same tiny surface as an openpyxl.Workbook (`.sheetnames`,
    `wb[name]`), backed by xlrd for legacy `.xls`."""

    def __init__(self, path):
        book = xlrd.open_workbook(path)
        self.sheetnames = book.sheet_names()
        self._sheets = {name: _XlsSheet(book.sheet_by_name(name), book.datemode)
                        for name in self.sheetnames}

    def __getitem__(self, name):
        return self._sheets[name]


def _open_workbook(path):
    """.xls (legacy BIFF) via xlrd; .xlsx/.xlsm (OOXML) via openpyxl —
    same reader surface either way, see module docstring."""
    if Path(path).suffix.lower() == ".xls":
        return _XlsWorkbook(path)
    return openpyxl.load_workbook(path, data_only=True)


def _norm_header(text):
    """Whitespace/case-tolerant header key, so 'CO6  No' or 'co6 no'
    still matches 'CO6 No'."""
    if text is None:
        return ""
    return re.sub(r"\s+", " ", str(text).replace("\xa0", " ")).strip().lower()


_HEADER_LOOKUP = {_norm_header(h): f for h, f in HEADER_TO_FIELD.items()}
_REQUIRED_FIELDS = tuple(HEADER_TO_FIELD[h] for h in REQUIRED_HEADERS)


class _Header(NamedTuple):
    """Where a sheet's Bill Status table starts and what its columns are.

    `row` is None when the sheet carries no such table — then `missing`
    names the REQUIRED headers absent from its closest candidate row, so
    a rejected upload can say what it was looking for.
    """
    row: Optional[int]
    cols: dict          # silver field name -> column index
    extras: dict        # unknown column's own header text -> column index
    missing: list       # REQUIRED headers absent from the closest row


def _find_header(ws, max_scan=MAX_HEADER_SCAN_ROWS) -> _Header:
    """The row (within the first few) that names every REQUIRED column."""
    best_missing: list = []
    for r in range(1, min(ws.max_row, max_scan) + 1):
        row = next(ws.iter_rows(min_row=r, max_row=r, values_only=True), ())
        cols, extras = {}, {}
        for idx, cell in enumerate(row):
            text = _norm_header(cell)
            if not text:
                continue
            field = _HEADER_LOOKUP.get(text)
            if field:
                cols.setdefault(field, idx)
            else:
                # a column this map doesn't know: keep it under its own
                # header text (whitespace-normalised, otherwise verbatim —
                # silver speaks the source's vocabulary), first occurrence
                # wins, never shadowing a known field name
                label = re.sub(r"\s+", " ",
                               str(cell).replace("\xa0", " ")).strip()
                if label not in FIELDS:
                    extras.setdefault(label, idx)
        missing = [h for h, f in zip(REQUIRED_HEADERS, _REQUIRED_FIELDS)
                   if f not in cols]
        if not missing:
            return _Header(r, cols, extras, [])
        # remember the closest near-miss for the error message
        if cols and (not best_missing or len(missing) < len(best_missing)):
            best_missing = missing
    return _Header(None, {}, {}, best_missing)


def parse_bill_status(path):
    """
    Parse every bill row in the workbook into one Silver record per bill.
    Accepts .xlsx/.xlsm (OOXML) and legacy .xls (BIFF) alike.

    Scans every worksheet for one whose header row names the REQUIRED
    columns (order and starting column don't matter); a worksheet without
    that header is skipped, not treated as data. Known optional columns
    that the export dropped come back all-NA, and columns it ADDED come
    back under their own header text. Raises BillStatusFormatError if no
    worksheet matches at all.
    """
    wb = _open_workbook(path)
    records = []
    matched_sheets = []
    near_miss: list = []
    # unknown columns in first-seen order, so the frame's column order is
    # deterministic across sheets that carry different extras
    extra_fields: list = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header = _find_header(ws)
        if header.row is None:
            # across sheets as within one: report the closest thing to a
            # Bill Status header the workbook had
            if header.missing and (not near_miss
                                   or len(header.missing) < len(near_miss)):
                near_miss = header.missing
            continue
        matched_sheets.append(sheet_name)
        for label in header.extras:
            if label not in extra_fields:
                extra_fields.append(label)
        read = {**header.cols, **header.extras}

        for r, row in enumerate(
                ws.iter_rows(min_row=header.row + 1, values_only=True),
                start=header.row + 1):
            if all(v is None for v in row):
                continue
            rec = {field: (row[idx] if idx < len(row) else None)
                  for field, idx in read.items()}
            rec["Sheet"] = sheet_name
            rec["DataRow"] = r
            records.append(rec)

    if not matched_sheets:
        raise BillStatusFormatError(
            "no Bill Status header row found in any worksheet; expected a "
            "table whose header names at least "
            + ", ".join(REQUIRED_HEADERS)
            + (f" (closest row was missing: {', '.join(near_miss)})"
               if near_miss else ""))

    df = pd.DataFrame(records, columns=FIELDS + extra_fields)
    if df.empty:
        return df

    # '----' -> null, identifier fields only (CO7No is the one place it's
    # documented to appear; harmless no-op elsewhere)
    for col in IDENTIFIER_FIELDS:
        placeholder = df[col].map(
            lambda v: isinstance(v, str)
            and bool(_DASH_PLACEHOLDER_RE.match(v.strip())))
        if placeholder.any():
            df[col] = df[col].astype(object)
            df.loc[placeholder, col] = None

    for col in AMOUNT_FIELDS:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    # dayfirst covers a future export writing dates as dd/mm/yyyy text;
    # a real Excel date cell (this export's normal case) passes through
    # unchanged. Any non-date placeholder ('NA', '-', '----', blank) the
    # export uses in a date column fails to parse and becomes NaT — one
    # rule instead of an enumerated placeholder list.
    for col in DATE_FIELDS:
        # format="mixed": a column may hold real datetime cells alongside
        # placeholder strings ('NA', '----') that must fail to a single
        # NaT each, not one shared format for the whole column
        df[col] = pd.to_datetime(df[col], errors="coerce", dayfirst=True,
                                 format="mixed")

    # Zone/PartyName/PartyCode/Status/AccountingUnit/ReasonForReturn/
    # RecoveryDetails and the identifier fields above are left exactly as
    # read — Silver is source-native, never transformed. So are the
    # unknown extra columns: nothing here knows what a column IREPS adds
    # next MEANS, and guessing a dtype for it would be an invention.
    return df
